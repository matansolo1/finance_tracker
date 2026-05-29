import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

def upgrade_excel():
    wb = openpyxl.load_workbook('expenses.xlsx')
    sheet = wb.active
    
    # 1. Colors & Fills
    navy_header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid') # Deep steel blue
    mint_income_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid') # Soft mint
    light_grey_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')  # Neutral light grey
    soft_blue_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')   # Soft blue-grey
    
    # Fonts
    font_header = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    font_bold_dark = Font(name='Segoe UI', size=11, bold=True, color='1F4E79')
    font_regular = Font(name='Segoe UI', size=11, bold=False, color='000000')
    font_bold = Font(name='Segoe UI', size=11, bold=True, color='000000')
    
    # Alignments
    align_right = Alignment(horizontal='right', vertical='center', wrap_text=True)
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Borders
    thin_side_grey = Side(border_style='thin', color='D9D9D9')
    thin_side_navy = Side(border_style='thin', color='1F4E79')
    double_side_black = Side(border_style='double', color='000000')
    
    border_all_thin = Border(left=thin_side_grey, right=thin_side_grey, top=thin_side_grey, bottom=thin_side_grey)
    border_header = Border(left=thin_side_grey, right=thin_side_grey, top=thin_side_navy, bottom=thin_side_navy)
    border_summary = Border(left=thin_side_grey, right=thin_side_grey, top=thin_side_navy, bottom=thin_side_navy)
    border_profit = Border(left=thin_side_grey, right=thin_side_grey, top=thin_side_navy, bottom=double_side_black)
    
    # 2. Process and style cell-by-cell
    max_row = sheet.max_row
    max_col = sheet.max_column
    
    # Enable gridlines explicitly
    sheet.views.sheetView[0].showGridLines = True
    
    # Currency Formatting String
    currency_format = '"₪"#,##0;[Red]"₪"(-#,##0);"-";@'
    
    # Map row classes based on known structure
    # Row 1: Header
    # Row 2: Income
    # Row 3 to 12: Expenses
    # Row 13: Summary Total Expenses
    # Row 14: Net disposable income
    
    for row_idx in range(1, max_row + 1):
        row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx))[0]
        
        # Determine the role of the row
        is_header = (row_idx == 1)
        is_income = (row_idx == 2)
        is_summary = (row_idx == 13 or (row[0].value == 'סיכום' if row[0].value else False))
        is_profit = (row_idx == 14 or (row[0].value == 'שורת רווח' if row[0].value else False))
        is_expense = not (is_header or is_income or is_summary or is_profit)
        
        for col_idx, cell in enumerate(row, start=1):
            val = cell.value
            
            # Handle 'NA' strings in numeric data columns (C, D, E)
            if col_idx in [3, 4, 5] and str(val).strip().upper() == 'NA':
                cell.value = 0 # Replace with 0 or None. Let's use 0 so summation continues seamlessly.
                val = 0
            
            # Base Font
            if is_header:
                cell.font = font_header
                cell.fill = navy_header_fill
                cell.alignment = align_center
                cell.border = border_header
            elif is_income:
                cell.font = font_bold if col_idx in [1, 2, 6] else font_regular
                cell.fill = mint_income_fill
                cell.alignment = align_right
                cell.border = border_all_thin
            elif is_summary:
                cell.font = font_bold_dark
                cell.fill = light_grey_fill
                cell.alignment = align_right
                cell.border = border_summary
            elif is_profit:
                cell.font = font_bold_dark
                cell.fill = soft_blue_fill
                cell.alignment = align_right
                cell.border = border_profit
            else: # regular expense
                cell.font = font_regular
                cell.alignment = align_right
                cell.border = border_all_thin
                
            # Number formatting for Columns C, D, E, F (value/formulas)
            if col_idx in [3, 4, 5, 6] and not is_header:
                # If it's a numeric value or a formula, apply currency formatting
                cell.number_format = currency_format

    # 3. Add Dropdown Category Data Validation to Column A (Categories)
    categories = ["הכנסות", "אשראי", "מגורים", "רכב", "סיכום", "שורת רווח"]
    dv = DataValidation(type="list", formula1=f'"{",".join(categories)}"', allow_blank=True)
    dv.error = 'הקטגוריה שהזנת אינה ברשימה המאושרת.'
    dv.errorTitle = 'קטגוריה שגויה'
    dv.prompt = 'אנא בחר קטגוריה מהרשימה.'
    dv.promptTitle = 'בחירת קטגוריה'
    
    sheet.add_data_validation(dv)
    # Apply to all cells in Column A from row 2 down to row 14
    dv.add(f"A2:A{max_row}")

    # 4. Freeze first row and columns A and B
    sheet.freeze_panes = 'C2'
    
    # 5. Auto-fit column widths
    for col in sheet.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            val_str = str(cell.value or '')
            # If it's a formula, provide a reasonable estimate of its length
            if val_str.startswith('='):
                val_str = "₪12,000" # Dummy estimate for formula length
            
            # Simple length count (Hebrew chars count as 1 here, but we can add some padding)
            if len(val_str) > max_len:
                max_len = len(val_str)
                
        # Set column width with padding
        sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Save upgraded workbook
    wb.save('expenses.xlsx')
    print("Excel file successfully upgraded!")

if __name__ == '__main__':
    upgrade_excel()
