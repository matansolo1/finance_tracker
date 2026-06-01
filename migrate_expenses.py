import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def migrate():
    # Load old expenses to verify or just build the dictionary directly
    # based on verified values:
    wb_old = openpyxl.load_workbook('expenses.xlsx')
    ws_old = wb_old.active
    
    # We will construct a new workbook
    wb_new = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb_new.active
    wb_new.remove(default_sheet)
    
    # Create the two sheets
    ws_actual = wb_new.create_sheet(title="תנועות_בפועל")
    ws_rules = wb_new.create_sheet(title="הגדרות_וחוקים")
    
    # Set RTL for Hebrew layout
    ws_actual.views.sheetView[0].showGridLines = True
    ws_rules.views.sheetView[0].showGridLines = True
    ws_actual.sheet_view.rightToLeft = True
    ws_rules.sheet_view.rightToLeft = True
    
    # Stylings
    navy_header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid') # Deep steel blue
    font_header = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    font_regular = Font(name='Segoe UI', size=11, color='000000')
    align_center = Alignment(horizontal='center', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    currency_format = '"₪"#,##0;[Red]"₪"(-#,##0);"-";@'
    
    # --- 1. POPULATE תנועות_בפועל ---
    headers_actual = ["תאריך", "קטגוריה", "סעיף", "סכום", "הערות"]
    ws_actual.append(headers_actual)
    
    # Style actual headers
    for col_idx in range(1, len(headers_actual) + 1):
        cell = ws_actual.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = navy_header_fill
        cell.alignment = align_center
        cell.border = thin_border
        
    actual_rows = [
        # Month 1 - Jan 2026
        ("01/01/2026", "הכנסות", "שכר נטו (ללא הטבה)", 12000, "שכר חודש ינואר"),
        ("01/01/2026", "אשראי", "הוצאות אשראי כלליות", 1500, "אשראי ינואר"),
        ("01/01/2026", "מגורים", "שכירות (שכ\"ד)", 2700, "שכירות ינואר"),
        ("01/01/2026", "מגורים", "חשמל (ממוצע)", 89, "חשמל ינואר"),
        ("01/01/2026", "מגורים", "גז (דוד גז - מקלחות)", 0, "גז ינואר"),
        ("01/01/2026", "מגורים", "מים (ממוצע לנפש)", 0, "מים ינואר"),
        ("01/01/2026", "רכב", "הלוואה לרכב", 1000, "הלוואה ינואר"),
        ("01/01/2026", "רכב", "ביטוח רכב (יחסי)", 583, "ביטוח ינואר"),
        ("01/01/2026", "רכב", "דלק לעבודה (תרדיון)", 120, "דלק ינואר"),
        ("01/01/2026", "רכב", "נסיעות לראשל\"צ (3)", 400, "נסיעות ינואר"),
        ("01/01/2026", "רכב", "החלפת קלאצ'", 1500, "קלאץ' תשלום 1 מתוך 7"),
        
        # Month 2 - Feb 2026
        ("01/02/2026", "הכנסות", "שכר נטו (ללא הטבה)", 12000, "שכר חודש פברואר"),
        ("01/02/2026", "אשראי", "הוצאות אשראי כלליות", 1500, "אשראי פברואר"),
        ("01/02/2026", "מגורים", "שכירות (שכ\"ד)", 2700, "שכירות פברואר"),
        ("01/02/2026", "מגורים", "חשמל (ממוצע)", 89, "חשמל פברואר"),
        ("01/02/2026", "מגורים", "גז (דוד גז - מקלחות)", 0, "גז פברואר"),
        ("01/02/2026", "מגורים", "מים (ממוצע לנפש)", 0, "מים פברואר"),
        ("01/02/2026", "רכב", "הלוואה לרכב", 1000, "הלוואה פברואר"),
        ("01/02/2026", "רכב", "ביטוח רכב (יחסי)", 583, "ביטוח פברואר"),
        ("01/02/2026", "רכב", "דלק לעבודה (תרדיון)", 120, "דלק פברואר"),
        ("01/02/2026", "רכב", "נסיעות לראשל\"צ (3)", 400, "נסיעות פברואר"),
        ("01/02/2026", "רכב", "החלפת קלאצ'", 1500, "קלאץ' תשלום 2 מתוך 7"),
        
        # Month 3 - Mar 2026
        ("01/03/2026", "הכנסות", "שכר נטו (ללא הטבה)", 12000, "שכר חודש מרץ"),
        ("01/03/2026", "אשראי", "הוצאות אשראי כלליות", 1500, "אשראי מרץ"),
        ("01/03/2026", "מגורים", "שכירות (שכ\"ד)", 2700, "שכירות מרץ"),
        ("01/03/2026", "מגורים", "חשמל (ממוצע)", 89, "חשמל מרץ"),
        ("01/03/2026", "מגורים", "גז (דוד גז - מקלחות)", 0, "גז מרץ"),
        ("01/03/2026", "מגורים", "מים (ממוצע לנפש)", 0, "מים מרץ"),
        ("01/03/2026", "רכב", "הלוואה לרכב", 1000, "הלוואה מרץ"),
        ("01/03/2026", "רכב", "ביטוח רכב (יחסי)", 584, "ביטוח מרץ"),
        ("01/03/2026", "רכב", "דלק לעבודה (תרדיון)", 120, "דלק מרץ"),
        ("01/03/2026", "רכב", "נסיעות לראשל\"צ (3)", 400, "נסיעות מרץ"),
        ("01/03/2026", "רכב", "החלפת קלאצ'", 1500, "קלאץ' תשלום 3 מתוך 7")
    ]
    
    for row_data in actual_rows:
        ws_actual.append(row_data)
        
    # Style actual cells
    for row_idx in range(2, ws_actual.max_row + 1):
        for col_idx in range(1, len(headers_actual) + 1):
            cell = ws_actual.cell(row=row_idx, column=col_idx)
            cell.font = font_regular
            cell.border = thin_border
            if col_idx in [1, 2, 3]:
                cell.alignment = align_center
            elif col_idx == 4:
                cell.alignment = align_right
                cell.number_format = currency_format
            else:
                cell.alignment = align_right

    # --- 2. POPULATE הגדרות_וחוקים ---
    headers_rules = ["סעיף", "קטגוריה", "סכום_דיפולט", "סוג_חוק", "תאריך_התחלה", "תאריך_סיום", "משולם_באשראי"]
    ws_rules.append(headers_rules)
    
    # Style rules headers
    for col_idx in range(1, len(headers_rules) + 1):
        cell = ws_rules.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = navy_header_fill
        cell.alignment = align_center
        cell.border = thin_border
        
    rules_rows = [
        ("שכר נטו (ללא הטבה)", "הכנסות", 12000, "Fixed Exact", "01/01/2026", None, False),
        ("הוצאות אשראי כלליות", "אשראי", 1500, "Fixed Estimate", "01/01/2026", None, False),
        ("שכירות (שכ\"ד)", "מגורים", 2700, "Fixed Exact", "01/01/2026", "01/01/2027", False),
        ("חשמל (ממוצע)", "מגורים", 89, "Fixed Estimate", "01/01/2026", None, True),
        ("גז (דוד גז - מקלחות)", "מגורים", 0, "Fixed Estimate", "01/01/2026", None, True),
        ("מים (ממוצע לנפש)", "מגורים", 0, "Fixed Estimate", "01/01/2026", None, True),
        ("הלוואה לרכב", "רכב", 1000, "Fixed Exact", "01/01/2026", "01/04/2028", False),
        ("ביטוח רכב (יחסי)", "רכב", 583, "Fixed Exact", "01/01/2026", None, True),
        ("דלק לעבודה (תרדיון)", "רכב", 120, "Fixed Estimate", "01/01/2026", None, True),
        ("נסיעות לראשל\"צ (3)", "רכב", 400, "Fixed Estimate", "01/01/2026", None, True),
        ("החלפת קלאצ'", "רכב", 1500, "Fixed Exact", "01/01/2026", "01/07/2026", True)
    ]
    
    for row_data in rules_rows:
        ws_rules.append(row_data)
        
    # Style rules cells
    for row_idx in range(2, ws_rules.max_row + 1):
        for col_idx in range(1, len(headers_rules) + 1):
            cell = ws_rules.cell(row=row_idx, column=col_idx)
            cell.font = font_regular
            cell.border = thin_border
            if col_idx in [1, 2, 4, 5, 6, 7]:
                cell.alignment = align_center
            elif col_idx == 3:
                cell.alignment = align_right
                cell.number_format = currency_format

    # --- 3. AUTO-FIT COLUMNS FOR BOTH SHEETS ---
    for ws in [ws_actual, ws_rules]:
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            max_len = 0
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Save to expenses.xlsx (overwrites old structure)
    import os
    try:
        wb_new.save('expenses.xlsx')
        print("Database successfully updated with two sheets: תנועות_בפועל and הגדרות_וחוקים!")
    except PermissionError:
        print("WARNING: Could not write directly to 'expenses.xlsx' because the file is open in another program (e.g., Excel).")
        print("Saving to 'expenses_new.xlsx' instead...")
        wb_new.save('expenses_new.xlsx')
        print("Successfully saved to 'expenses_new.xlsx'!")
        print("Please close 'expenses.xlsx' in Excel and manually replace it with 'expenses_new.xlsx' or run 'python migrate_expenses.py' again.")

if __name__ == '__main__':
    migrate()
