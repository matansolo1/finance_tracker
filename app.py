import os
from dotenv import load_dotenv
import shutil
import urllib.parse
from functools import wraps
from datetime import datetime, date
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask import Flask, jsonify, request, send_from_directory, session, redirect, url_for, send_file, flash, render_template

app = Flask(__name__)
# Secure Flask session configuration
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "super-secret-key-for-fincontrol-saas")
# טעינת משתני הסביבה מקובץ .env
load_dotenv()
# --- GOOGLE OAUTH CONFIGURATION ---
GOOGLE_CLIENT_ID = os.environ.get("GOOGLE_CLIENT_ID")
GOOGLE_CLIENT_SECRET = os.environ.get("GOOGLE_CLIENT_SECRET")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

def get_user_excel_file():
    """Resolves the Excel file path dynamically using the active user's session identifier."""
    user_google_id = session.get('user_google_id')
    if not user_google_id:
        return None
    return os.path.join(BASE_DIR, 'users_data', user_google_id, 'expenses.xlsx')

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_google_id' not in session:
            if request.path.startswith('/api/'):
                return jsonify({'error': 'Unauthorized'}), 401
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def parse_date(val):
    if isinstance(val, datetime):
        return val
    if isinstance(val, date):
        return datetime(val.year, val.month, val.day)
    if isinstance(val, (int, float)):
        try:
            from datetime import timedelta
            # Excel date starts from 1900-01-01. Excel leap year bug offset is 1899-12-30.
            if val >= 61:
                return datetime(1899, 12, 30) + timedelta(days=val)
            else:
                return datetime(1899, 12, 31) + timedelta(days=val)
        except Exception:
            pass
    if isinstance(val, str):
        val = val.strip()
        # Support full dates, as well as month-only date patterns (e.g. 5/2026, 05/2026)
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y", "%m/%Y", "%Y-%m", "%m/%y"):
            try:
                return datetime.strptime(val, fmt)
            except ValueError:
                pass
    return None

def format_date(dt):
    if not dt:
        return ""
    return dt.strftime("%d/%m/%Y")

def get_sheet_data(sheet_name, excel_file):
    """Safely loads rows from a given sheet and handles basic cell reading."""
    if not excel_file or not os.path.exists(excel_file):
        return []
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []
    sheet = wb[sheet_name]
    rows = []
    for row in sheet.iter_rows(values_only=True):
        rows.append(list(row))
    wb.close()
    return rows

def apply_row_styles(sheet, r_idx, is_header=False):
    """Utility to style newly created or updated expense/rule rows."""
    font_header = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    font_regular = Font(name='Segoe UI', size=11, bold=False, color='000000')
    
    navy_header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    align_center = Alignment(horizontal='center', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    row_cells = list(sheet.iter_rows(min_row=r_idx, max_row=r_idx))[0]
    
    for col_idx, cell in enumerate(row_cells, start=1):
        if is_header:
            cell.font = font_header
            cell.fill = navy_header_fill
            cell.alignment = align_center
            cell.border = thin_border
        else:
            cell.font = font_regular
            cell.border = thin_border
            cell.alignment = align_right
            
    # Auto-fit columns
    for col in sheet.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

def get_credit_card_deductions(target_year, target_month, excel_file):
    """
    Returns:
      total_deduction: float (sum of flagged active rules + flagged actual transactions for that month)
      breakdown: list of dicts with details of what was deducted
    """
    actuals_raw = get_sheet_data("תנועות_בפועל", excel_file)
    rules_raw = get_sheet_data("הגדרות_וחוקים", excel_file)
    
    # Map actuals of this month by (item, category) -> (amount, is_credit_card)
    actual_amounts = {}
    actual_credit_card_flags = {}
    if len(actuals_raw) > 1:
        for r in actuals_raw[1:]:
            if not any(val is not None for val in r):
                continue
            dt_parsed = parse_date(r[0])
            if dt_parsed and dt_parsed.year == target_year and dt_parsed.month == target_month:
                item = str(r[2] or "").strip()
                category = str(r[1] or "").strip()
                actual_amounts[(item, category)] = float(r[3] or 0)
                # Column 6 (index 5): is_credit_card flag on actual transaction
                is_cc = False
                if len(r) > 5 and r[5] is not None:
                    is_cc = (str(r[5]).strip().upper() == 'TRUE' or r[5] is True)
                actual_credit_card_flags[(item, category)] = is_cc

    deductions_sum = 0.0
    deductions_list = []
    # Track which (item, category) pairs have already been added (to avoid double-counting)
    deducted_keys = set()

    # Step 1: Process rules with is_credit_card=TRUE (existing logic)
    if len(rules_raw) > 1:
        for r in rules_raw[1:]:
            if not any(val is not None for val in r):
                continue
            item = str(r[0] or "").strip()
            category = str(r[1] or "").strip()
            default_amount = float(r[2] or 0)
            start_dt = parse_date(r[4])
            end_dt = parse_date(r[5]) if r[5] else None
            
            is_credit_card = False
            if len(r) > 6 and r[6] is not None:
                is_credit_card = (str(r[6]).strip().upper() == 'TRUE' or r[6] is True)
                
            if is_credit_card and start_dt:
                req_dt_compare = date(target_year, target_month, 1)
                start_compare = date(start_dt.year, start_dt.month, 1)
                end_compare = date(end_dt.year, end_dt.month, 1) if end_dt else None
                
                is_active = (req_dt_compare >= start_compare) and (end_compare is None or req_dt_compare <= end_compare)
                if is_active:
                    key = (item, category)
                    # Use actual amount if exists, else default_amount
                    amt = actual_amounts.get(key, default_amount)
                    deductions_sum += amt
                    deductions_list.append({
                        "item": item,
                        "category": category,
                        "amount": amt,
                        "source": "rule"
                    })
                    deducted_keys.add(key)

    # Step 2: Process actual transactions flagged as is_credit_card=TRUE (new logic)
    # Only add those that don't already have a matching rule deduction
    for (item, category), is_cc in actual_credit_card_flags.items():
        if is_cc and (item, category) not in deducted_keys:
            amt = actual_amounts.get((item, category), 0.0)
            deductions_sum += amt
            deductions_list.append({
                "item": item,
                "category": category,
                "amount": amt,
                "source": "actual"
            })
            deducted_keys.add((item, category))
                    
    return deductions_sum, deductions_list

def get_monthly_calculations(target_year, target_month, excel_file):
    """
    Retrieves, parses, merges and aggregates transactions for a given month & year.
    Returns: (list of merged transactions, summary dict)
    """
    actuals_raw = get_sheet_data("תנועות_בפועל", excel_file)
    rules_raw = get_sheet_data("הגדרות_וחוקים", excel_file)
    
    actual_txs = []
    if len(actuals_raw) > 1:
        for r in actuals_raw[1:]:
            if not any(val is not None for val in r):
                continue
            dt_parsed = parse_date(r[0])
            if dt_parsed and dt_parsed.year == target_year and dt_parsed.month == target_month:
                # Column 6 (index 5): is_credit_card flag on actual transaction
                is_cc = False
                if len(r) > 5 and r[5] is not None:
                    is_cc = (str(r[5]).strip().upper() == 'TRUE' or r[5] is True)
                actual_txs.append({
                    "date": format_date(dt_parsed),
                    "category": str(r[1] or "").strip(),
                    "item": str(r[2] or "").strip(),
                    "amount": float(r[3] or 0),
                    "notes": str(r[4] or "").strip(),
                    "is_credit_card": is_cc,
                    "is_estimate": False,
                    "status": "actual"
                })
                
    active_rules = []
    if len(rules_raw) > 1:
        for r in rules_raw[1:]:
            if not any(val is not None for val in r):
                continue
            item = str(r[0] or "").strip()
            category = str(r[1] or "").strip()
            default_amount = float(r[2] or 0)
            rule_type = str(r[3] or "").strip()
            start_dt = parse_date(r[4])
            end_dt = parse_date(r[5]) if r[5] else None
            
            is_credit_card = False
            if len(r) > 6 and r[6] is not None:
                is_credit_card = (str(r[6]).strip().upper() == 'TRUE' or r[6] is True)
            
            if start_dt:
                req_dt_compare = date(target_year, target_month, 1)
                start_compare = date(start_dt.year, start_dt.month, 1)
                end_compare = date(end_dt.year, end_dt.month, 1) if end_dt else None
                
                is_active = (req_dt_compare >= start_compare) and (end_compare is None or req_dt_compare <= end_compare)
                if is_active:
                    active_rules.append({
                        "item": item,
                        "category": category,
                        "default_amount": default_amount,
                        "rule_type": rule_type,
                        "start_date": format_date(start_dt),
                        "end_date": format_date(end_dt) if end_dt else "",
                        "is_credit_card": is_credit_card
                    })

    # Merge & Override
    merged_txs = list(actual_txs)
    actual_items = { (tx["item"], tx["category"]) for tx in actual_txs }
    
    for rule in active_rules:
        rule_key = (rule["item"], rule["category"])
        if rule_key not in actual_items:
            merged_txs.append({
                "date": f"01/{target_month:02d}/{target_year}",
                "category": rule["category"],
                "item": rule["item"],
                "amount": rule["default_amount"],
                "notes": "הערכה דינמית",
                "is_estimate": True,
                "status": "estimated"
            })
            
    # Aggregations
    total_income = 0.0
    fixed_expenses_total = 0.0
    variable_expenses_total = 0.0
    
    for tx in merged_txs:
        cat = tx["category"]
        amt = tx["amount"]
        item = tx["item"]
        if cat == "הכנסות":
            total_income += amt
        else:
            if item == "הוצאות אשראי כלליות":
                variable_expenses_total += amt
            else:
                fixed_expenses_total += amt
                
    net_savings = total_income - (fixed_expenses_total + variable_expenses_total)
    
    summary = {
        "total_income": total_income,
        "fixed_expenses": fixed_expenses_total,
        "variable_expenses": variable_expenses_total,
        "fixed_expenses_total": fixed_expenses_total,
        "variable_expenses_total": variable_expenses_total,
        "net_savings": net_savings
    }
    
    return merged_txs, summary

# --- OAUTH & ROUTING ---

@app.route('/login')
def login():
    print("--> DEBUG: Login route '/login' was hit!", flush=True)
    print(f"--> DEBUG: Current session keys: {list(session.keys())}", flush=True)
    if 'user_google_id' in session:
        return redirect(url_for('index'))
    return render_template('login.html')

@app.route('/login/google')
def login_google():
    auth_url = "https://accounts.google.com/o/oauth2/v2/auth"
    scheme = request.headers.get('X-Forwarded-Proto', request.scheme)
    redirect_uri = url_for('oauth_callback', _external=True, _scheme=scheme)
    
    params = {
        "client_id": GOOGLE_CLIENT_ID,
        "redirect_uri": redirect_uri,
        "response_type": "code",
        "scope": "openid email profile",
        "access_type": "offline",
        "prompt": "consent"
    }
    target_url = f"{auth_url}?{urllib.parse.urlencode(params)}"
    return redirect(target_url)

@app.route('/login/callback')
def oauth_callback():
    code = request.args.get('code')
    if not code:
        return "Missing authorization code", 400
        
    scheme = request.headers.get('X-Forwarded-Proto', request.scheme)
    redirect_uri = url_for('oauth_callback', _external=True, _scheme=scheme)
    
    token_url = "https://oauth2.googleapis.com/token"
    token_data = {
        "code": code,
        "client_id": GOOGLE_CLIENT_ID,
        "client_secret": GOOGLE_CLIENT_SECRET,
        "redirect_uri": redirect_uri,
        "grant_type": "authorization_code"
    }
    
    token_response = requests.post(token_url, data=token_data)
    if token_response.status_code != 200:
        return f"Failed to acquire token: {token_response.text}", 400
        
    token_json = token_response.json()
    access_token = token_json.get("access_token")
    
    userinfo_url = "https://www.googleapis.com/oauth2/v3/userinfo"
    userinfo_headers = {"Authorization": f"Bearer {access_token}"}
    userinfo_response = requests.get(userinfo_url, headers=userinfo_headers)
    
    if userinfo_response.status_code != 200:
        return f"Failed to retrieve user info: {userinfo_response.text}", 400
        
    user_info = userinfo_response.json()
    user_google_id = user_info.get("sub")
    
    if not user_google_id:
        return "Failed to extract user ID", 400
        
    # Set secure session details
    session['user_google_id'] = user_google_id
    session['user_email'] = user_info.get("email")
    session['user_name'] = user_info.get("name")
    session.permanent = True
    
    # Dynamic path handling & Automated directory initialization
    user_dir = os.path.join(BASE_DIR, 'users_data', user_google_id)
    if not os.path.exists(user_dir):
        os.makedirs(user_dir, exist_ok=True)
        template_file = os.path.join(BASE_DIR, 'users_data', 'template_empty.xlsx')
        user_excel_file = os.path.join(user_dir, 'expenses.xlsx')
        if os.path.exists(template_file):
            shutil.copy(template_file, user_excel_file)
        session['is_new_user'] = True
            
    return redirect(url_for('index'))

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/')
def index():
    print("--> DEBUG: Root route '/' was hit!", flush=True)
    print(f"--> DEBUG: Current session keys: {list(session.keys())}", flush=True)
    if 'user_google_id' not in session:
        return redirect(url_for('login'))
    return render_template('index.html')

@app.route('/download-excel')
@login_required
def download_excel():
    excel_file = get_user_excel_file()
    if not excel_file or not os.path.exists(excel_file):
        user_google_id = session.get('user_google_id')
        user_dir = os.path.join(BASE_DIR, 'users_data', user_google_id)
        os.makedirs(user_dir, exist_ok=True)
        template_file = os.path.join(BASE_DIR, 'users_data', 'template_empty.xlsx')
        user_excel_file = os.path.join(user_dir, 'expenses.xlsx')
        if os.path.exists(template_file):
            shutil.copy(template_file, user_excel_file)
            
    if not excel_file or not os.path.exists(excel_file):
        return "קובץ האקסל לא נמצא.", 404
        
    return send_file(excel_file, as_attachment=True, download_name='my_expenses_backup.xlsx')

@app.route('/download-template')
@login_required
def download_template():
    template_file = os.path.join(BASE_DIR, 'users_data', 'template_empty.xlsx')
    if not os.path.exists(template_file):
        return "קובץ התבנית הריק לא נמצא.", 404
    return send_file(template_file, as_attachment=True, download_name='template_empty.xlsx')

@app.route('/upload-excel', methods=['POST'])
@login_required
def upload_excel():
    user_google_id = session.get('user_google_id')
    if not user_google_id:
        return redirect(url_for('login'))
        
    if 'file' not in request.files:
        flash('לא נבחר קובץ להעלאה.', 'error')
        return redirect(url_for('index', upload_error='לא נבחר קובץ להעלאה.'))
        
    file = request.files['file']
    if file.filename == '':
        flash('לא נבחר קובץ להעלאה.', 'error')
        return redirect(url_for('index', upload_error='לא נבחר קובץ להעלאה.'))
        
    if not file.filename.endswith('.xlsx'):
        flash('קובץ לא חוקי. אנא העלה קובץ אקסל בסיומת .xlsx בלבד.', 'error')
        return redirect(url_for('index', upload_error='סיומת קובץ לא חוקית. יש להעלות קובץ .xlsx בלבד.'))
        
    user_dir = os.path.join(BASE_DIR, 'users_data', user_google_id)
    os.makedirs(user_dir, exist_ok=True)
    temp_path = os.path.join(user_dir, 'temp_upload.xlsx')
    
    try:
        file.save(temp_path)
        
        # Validation using openpyxl
        wb = openpyxl.load_workbook(temp_path, read_only=True)
        sheets = wb.sheetnames
        wb.close()
        
        required_sheets = ['תנועות_בפועל', 'הגדרות_וחוקים']
        for r_sheet in required_sheets:
            if r_sheet not in sheets:
                if os.path.exists(temp_path):
                    os.remove(temp_path)
                msg = f"מבנה הגיליונות בקובץ אינו תקין. הקובץ חייב להכיל את הגיליונות: {', '.join(required_sheets)}"
                flash(msg, 'error')
                return redirect(url_for('index', upload_error=msg))
                
        # Overwrite existing file
        user_excel_file = os.path.join(user_dir, 'expenses.xlsx')
        if os.path.exists(user_excel_file):
            os.remove(user_excel_file)
        os.rename(temp_path, user_excel_file)
        
        flash('הקובץ הועלה ועודכן בהצלחה!', 'success')
        return redirect(url_for('index', upload_success='1'))
        
    except Exception as e:
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except Exception:
                pass
        msg = f"שגיאה בעיבוד הקובץ: {str(e)}"
        flash(msg, 'error')
        return redirect(url_for('index', upload_error=msg))

@app.route('/api/data', methods=['GET'])
@login_required
def get_data():
    try:
        excel_file = get_user_excel_file()
        if not excel_file or not os.path.exists(excel_file):
            return jsonify({'error': 'Active database not found. Please log in again.'}), 404
            
        year = int(request.args.get('year', 2026))
        month = int(request.args.get('month', 1))
        
        merged_transactions, month_summary = get_monthly_calculations(year, month, excel_file)
        
        if 1 <= month <= 3:
            quarter_months = [1, 2, 3]
        elif 4 <= month <= 6:
            quarter_months = [4, 5, 6]
        elif 7 <= month <= 9:
            quarter_months = [7, 8, 9]
        else:
            quarter_months = [10, 11, 12]
            
        q_income = q_fixed = q_variable = q_net = 0.0
        q_breakdown = []
        for qm in quarter_months:
            _, qm_summary = get_monthly_calculations(year, qm, excel_file)
            q_income += qm_summary["total_income"]
            q_fixed += qm_summary["fixed_expenses"]
            q_variable += qm_summary["variable_expenses"]
            q_net += qm_summary["net_savings"]
            q_breakdown.append({
                "month": qm,
                "summary": qm_summary
            })
            
        quarter_summary = {
            "total_income": q_income,
            "fixed_expenses": q_fixed,
            "variable_expenses": q_variable,
            "net_savings": q_net,
            "breakdown": q_breakdown
        }
        
        ytd_income = ytd_fixed = ytd_variable = ytd_net = 0.0
        ytd_breakdown = []
        for m in range(1, month + 1):
            _, m_summary = get_monthly_calculations(year, m, excel_file)
            ytd_income += m_summary["total_income"]
            ytd_fixed += m_summary["fixed_expenses"]
            ytd_variable += m_summary["variable_expenses"]
            ytd_net += m_summary["net_savings"]
            ytd_breakdown.append({
                "month": m,
                "summary": m_summary
            })
            
        year_summary = {
            "total_income": ytd_income,
            "fixed_expenses": ytd_fixed,
            "variable_expenses": ytd_variable,
            "net_savings": ytd_net,
            "breakdown": ytd_breakdown
        }
        
        deductions_sum, deductions_list = get_credit_card_deductions(year, month, excel_file)
        
        is_new_user = session.get('is_new_user', False)
        if is_new_user:
            session['is_new_user'] = False

        return jsonify({
            "transactions": merged_transactions,
            "month_summary": month_summary,
            "quarter_summary": quarter_summary,
            "year_summary": year_summary,
            "fixed_expenses_total": month_summary["fixed_expenses_total"],
            "variable_expenses_total": month_summary["variable_expenses_total"],
            "credit_card_deductions": {
                "total_deducted": deductions_sum,
                "breakdown": deductions_list
            },
            "user_email": session.get('user_email', ''),
            "user_name": session.get('user_name', ''),
            "is_new_user": is_new_user
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/expenses', methods=['POST'])
@login_required
def add_expense():
    try:
        excel_file = get_user_excel_file()
        if not excel_file or not os.path.exists(excel_file):
            return jsonify({'error': 'Active database not found. Please log in again.'}), 404
            
        req_data = request.json
        category = req_data.get('category')
        category = category.strip() if isinstance(category, str) else ""
        item = req_data.get('item')
        item = item.strip() if isinstance(item, str) else ""
        amount = float(req_data.get('amount') or 0)
        notes = req_data.get('notes')
        notes = notes.strip() if isinstance(notes, str) else ""
        
        year = req_data.get('year')
        month = req_data.get('month')
        date_str = req_data.get('date')
        date_str = date_str.strip() if isinstance(date_str, str) else ""
        
        if year and month:
            target_year = int(year)
            target_month = int(month)
            formatted_dt_str = f"01/{target_month:02d}/{target_year}"
        elif date_str:
            dt = parse_date(date_str)
            if not dt:
                return jsonify({'error': 'Invalid date format'}), 400
            target_year = dt.year
            target_month = dt.month
            formatted_dt_str = format_date(dt)
        else:
            return jsonify({'error': 'Year/Month or Date is required'}), 400
            
        if not category or not item:
            return jsonify({'error': 'Category and item are required'}), 400
            
        if item == "הוצאות אשראי כלליות":
            deductions_sum, _ = get_credit_card_deductions(target_year, target_month, excel_file)
            net_amount = amount - deductions_sum
            amount = max(0.0, net_amount)
            
        wb = openpyxl.load_workbook(excel_file)
        if "תנועות_בפועל" not in wb.sheetnames:
            wb.close()
            return jsonify({'error': 'Sheet תנועות_בפועל not found'}), 500
            
        sheet = wb["תנועות_בפועל"]
        
        found_row_idx = -1
        for r_idx in range(2, sheet.max_row + 1):
            cell_date_val = sheet.cell(row=r_idx, column=1).value
            cell_cat_val = sheet.cell(row=r_idx, column=2).value
            cell_item_val = sheet.cell(row=r_idx, column=3).value
            
            if cell_date_val is None or cell_cat_val is None or cell_item_val is None:
                continue
                
            dt_parsed = parse_date(cell_date_val)
            if dt_parsed and dt_parsed.year == target_year and dt_parsed.month == target_month:
                if str(cell_cat_val).strip() == category and str(cell_item_val).strip() == item:
                    found_row_idx = r_idx
                    break
                    
        is_credit_card = bool(req_data.get('is_credit_card', False))

        if found_row_idx != -1:
            sheet.cell(row=found_row_idx, column=4, value=amount)
            if notes:
                sheet.cell(row=found_row_idx, column=5, value=notes)
            sheet.cell(row=found_row_idx, column=6, value=is_credit_card)
            apply_row_styles(sheet, found_row_idx)
            sheet.cell(row=found_row_idx, column=4).number_format = '"₪"#,##0;[Red]"₪"(-#,##0);"-";@'
            msg = 'Transaction updated successfully'
        else:
            new_row_idx = sheet.max_row + 1
            sheet.cell(row=new_row_idx, column=1, value=formatted_dt_str)
            sheet.cell(row=new_row_idx, column=2, value=category)
            sheet.cell(row=new_row_idx, column=3, value=item)
            sheet.cell(row=new_row_idx, column=4, value=amount)
            sheet.cell(row=new_row_idx, column=5, value=notes or "עדכון ידני")
            sheet.cell(row=new_row_idx, column=6, value=is_credit_card)
            apply_row_styles(sheet, new_row_idx)
            sheet.cell(row=new_row_idx, column=4).number_format = '"₪"#,##0;[Red]"₪"(-#,##0);"-";@'
            msg = 'Transaction added successfully'
            
        try:
            wb.save(excel_file)
            wb.close()
        except PermissionError:
            wb.close()
            return jsonify({
                'success': False,
                'error': 'קובץ ה-Excel פתוח כרגע בתוכנה אחרת. אנא סגור אותו באקסל ונסה שוב כדי שהשינויים יישמרו.'
            }), 409
        
        return jsonify({'success': True, 'message': msg})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/categories', methods=['GET'])
@login_required
def get_categories():
    """Returns a unique list of all categories from both sheets."""
    try:
        excel_file = get_user_excel_file()
        if not excel_file or not os.path.exists(excel_file):
            return jsonify({'error': 'Active database not found. Please log in again.'}), 404
            
        categories = set()
        
        # Scan actual transactions
        actuals_raw = get_sheet_data("תנועות_בפועל", excel_file)
        if len(actuals_raw) > 1:
            for r in actuals_raw[1:]:
                if r and len(r) > 1 and r[1]:
                    cat = str(r[1]).strip()
                    if cat:
                        categories.add(cat)
        
        # Scan rules
        rules_raw = get_sheet_data("הגדרות_וחוקים", excel_file)
        if len(rules_raw) > 1:
            for r in rules_raw[1:]:
                if r and len(r) > 1 and r[1]:
                    cat = str(r[1]).strip()
                    if cat:
                        categories.add(cat)
        
        # Return sorted list
        return jsonify({'categories': sorted(list(categories))})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/analytics/trends', methods=['GET'])
@login_required
def get_category_trends():
    """
    Returns category spending trends across multiple periods.
    Query params: category (required), period_type (quarterly/annual/multi_year), 
                  year (required), month (for quarterly context)
    """
    try:
        excel_file = get_user_excel_file()
        if not excel_file or not os.path.exists(excel_file):
            return jsonify({'error': 'Active database not found. Please log in again.'}), 404
            
        category = request.args.get('category')
        if not category:
            return jsonify({'error': 'Category parameter is required'}), 400
            
        period_type = request.args.get('period_type', 'quarterly')  # quarterly, annual, multi_year
        year = int(request.args.get('year', 2026))
        month = int(request.args.get('month', 1))
        
        monthly_data = []
        
        if period_type == 'quarterly':
            # Determine quarter
            if 1 <= month <= 3:
                months = [1, 2, 3]
            elif 4 <= month <= 6:
                months = [4, 5, 6]
            elif 7 <= month <= 9:
                months = [7, 8, 9]
            else:
                months = [10, 11, 12]
                
            for m in months:
                amount = get_category_amount_for_month(category, year, m, excel_file)
                monthly_data.append({
                    'year': year,
                    'month': m,
                    'amount': amount
                })
                
        elif period_type == 'annual':
            # Full year
            for m in range(1, 13):
                amount = get_category_amount_for_month(category, year, m, excel_file)
                monthly_data.append({
                    'year': year,
                    'month': m,
                    'amount': amount
                })
                
        elif period_type == 'multi_year':
            # Last 3 years
            for y in range(year - 2, year + 1):
                for m in range(1, 13):
                    amount = get_category_amount_for_month(category, y, m, excel_file)
                    monthly_data.append({
                        'year': y,
                        'month': m,
                        'amount': amount
                    })
        
        # Calculate average
        total_amount = sum(item['amount'] for item in monthly_data)
        month_count = len(monthly_data)
        average = total_amount / month_count if month_count > 0 else 0.0
        
        return jsonify({
            'category': category,
            'period_type': period_type,
            'monthly_data': monthly_data,
            'total_amount': total_amount,
            'average_per_month': average,
            'month_count': month_count
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def get_category_amount_for_month(category, target_year, target_month, excel_file):
    """
    Helper function to calculate total spending for a specific category in a given month.
    For credit card category, applies the deduction logic.
    """
    merged_txs, _ = get_monthly_calculations(target_year, target_month, excel_file)
    
    total = 0.0
    for tx in merged_txs:
        if tx['category'] == category:
            total += tx['amount']
    
    return total

@app.route('/api/rules', methods=['POST'])
@login_required
def handle_rule():
    try:
        excel_file = get_user_excel_file()
        if not excel_file or not os.path.exists(excel_file):
            return jsonify({'error': 'Active database not found. Please log in again.'}), 404
            
        req_data = request.json
        item = req_data.get('item')
        item = item.strip() if isinstance(item, str) else ""
        category = req_data.get('category')
        category = category.strip() if isinstance(category, str) else ""
        amount = float(req_data.get('amount') or 0)
        rule_type = req_data.get('rule_type')
        rule_type = rule_type.strip() if isinstance(rule_type, str) else "Fixed Exact"
        start_date = req_data.get('start_date')
        start_date = start_date.strip() if isinstance(start_date, str) else ""
        end_date = req_data.get('end_date')
        end_date = end_date.strip() if isinstance(end_date, str) else ""
        is_credit_card = bool(req_data.get('is_credit_card', False))
        
        if not item:
            return jsonify({'error': 'Item is required'}), 400
            
        start_dt = parse_date(start_date) if start_date else datetime.now()
        end_dt = parse_date(end_date) if end_date else None
        
        formatted_start = format_date(start_dt)
        formatted_end = format_date(end_dt) if end_dt else ""
        
        wb = openpyxl.load_workbook(excel_file)
        if "הגדרות_וחוקים" not in wb.sheetnames:
            wb.close()
            return jsonify({'error': 'Sheet הגדרות_וחוקים not found'}), 500
            
        sheet = wb["הגדרות_וחוקים"]
        
        found_row_idx = -1
        for r_idx in range(2, sheet.max_row + 1):
            val = sheet.cell(row=r_idx, column=1).value
            if val and str(val).strip() == item:
                found_row_idx = r_idx
                break
                
        if found_row_idx != -1:
            if category:
                sheet.cell(row=found_row_idx, column=2, value=category)
            sheet.cell(row=found_row_idx, column=3, value=amount)
            if rule_type:
                sheet.cell(row=found_row_idx, column=4, value=rule_type)
            sheet.cell(row=found_row_idx, column=5, value=formatted_start)
            sheet.cell(row=found_row_idx, column=6, value=formatted_end)
            sheet.cell(row=found_row_idx, column=7, value=is_credit_card)
            apply_row_styles(sheet, found_row_idx)
            sheet.cell(row=found_row_idx, column=3).number_format = '"₪"#,##0;[Red]"₪"(-#,##0);"-";@'
            msg = 'Rule updated successfully'
        else:
            if not category:
                category = "שונות"
            new_row_idx = sheet.max_row + 1
            sheet.cell(row=new_row_idx, column=1, value=item)
            sheet.cell(row=new_row_idx, column=2, value=category)
            sheet.cell(row=new_row_idx, column=3, value=amount)
            sheet.cell(row=new_row_idx, column=4, value=rule_type)
            sheet.cell(row=new_row_idx, column=5, value=formatted_start)
            sheet.cell(row=new_row_idx, column=6, value=formatted_end)
            sheet.cell(row=new_row_idx, column=7, value=is_credit_card)
            apply_row_styles(sheet, new_row_idx)
            sheet.cell(row=new_row_idx, column=3).number_format = '"₪"#,##0;[Red]"₪"(-#,##0);"-";@'
            msg = 'Rule created successfully'
            
        try:
            wb.save(excel_file)
            wb.close()
        except PermissionError:
            wb.close()
            return jsonify({
                'success': False,
                'error': 'קובץ ה-Excel פתוח כרגע בתוכנה אחרת. אנא סגור אותו באקסל ונסה שוב כדי שהשינויים יישמרו.'
            }), 409
        
        return jsonify({'success': True, 'message': msg})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/expenses', methods=['PUT'])
@login_required
def edit_expense():
    """Edit an existing actual transaction by (date, category, item) key."""
    try:
        excel_file = get_user_excel_file()
        if not excel_file or not os.path.exists(excel_file):
            return jsonify({'error': 'Active database not found. Please log in again.'}), 404

        req_data = request.json
        category = req_data.get('category', '').strip()
        item = req_data.get('item', '').strip()
        amount = float(req_data.get('amount') or 0)
        notes = req_data.get('notes', '').strip()
        is_credit_card = bool(req_data.get('is_credit_card', False))

        # Date resolution
        date_str = req_data.get('date', '').strip()
        year = req_data.get('year')
        month = req_data.get('month')

        if year and month:
            target_year = int(year)
            target_month = int(month)
            formatted_dt_str = f"01/{target_month:02d}/{target_year}"
        elif date_str:
            dt = parse_date(date_str)
            if not dt:
                return jsonify({'error': 'Invalid date format'}), 400
            target_year = dt.year
            target_month = dt.month
            formatted_dt_str = format_date(dt)
        else:
            return jsonify({'error': 'Year/Month or Date is required'}), 400

        if not category or not item:
            return jsonify({'error': 'Category and item are required'}), 400

        wb = openpyxl.load_workbook(excel_file)
        if "תנועות_בפועל" not in wb.sheetnames:
            wb.close()
            return jsonify({'error': 'Sheet תנועות_בפועל not found'}), 500

        sheet = wb["תנועות_בפועל"]

        found_row_idx = -1
        for r_idx in range(2, sheet.max_row + 1):
            cell_date_val = sheet.cell(row=r_idx, column=1).value
            cell_cat_val = sheet.cell(row=r_idx, column=2).value
            cell_item_val = sheet.cell(row=r_idx, column=3).value

            if cell_date_val is None or cell_cat_val is None or cell_item_val is None:
                continue

            dt_parsed = parse_date(cell_date_val)
            if dt_parsed and dt_parsed.year == target_year and dt_parsed.month == target_month:
                if str(cell_cat_val).strip() == category and str(cell_item_val).strip() == item:
                    found_row_idx = r_idx
                    break

        if found_row_idx == -1:
            wb.close()
            return jsonify({'error': 'Transaction not found'}), 404

        sheet.cell(row=found_row_idx, column=1, value=formatted_dt_str)
        sheet.cell(row=found_row_idx, column=2, value=category)
        sheet.cell(row=found_row_idx, column=3, value=item)
        sheet.cell(row=found_row_idx, column=4, value=amount)
        sheet.cell(row=found_row_idx, column=5, value=notes or "עדכון ידני")
        sheet.cell(row=found_row_idx, column=6, value=is_credit_card)
        apply_row_styles(sheet, found_row_idx)
        sheet.cell(row=found_row_idx, column=4).number_format = '"₪"#,##0;[Red]"₪"(-#,##0);"-";@'

        try:
            wb.save(excel_file)
            wb.close()
        except PermissionError:
            wb.close()
            return jsonify({
                'success': False,
                'error': 'קובץ ה-Excel פתוח כרגע בתוכנה אחרת. אנא סגור אותו באקסל ונסה שוב כדי שהשינויים יישמרו.'
            }), 409

        return jsonify({'success': True, 'message': 'Transaction edited successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    # Dynamic Port Binding
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
